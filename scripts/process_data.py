#!/home/alex/.var/app/org.jupyter.JupyterLab/config/jupyterlab-desktop/jlab_server/bin/python3
"""
process_data.py
Joins Oregon county boundaries, library directory, graduation rates,
and dropout rates into a single GeoJSON for the portfolio map.

Data sources:
  Oregon_Counties_map_20260325.geojson       — county boundary polygons
  Oregon_Library_Directory_20260325.geojson  — library point features
  cohortmediafile2024-2025.xlsx              — ODE 4-year cohort graduation rates
  dropouttables2024-2025.xlsx                — ODE dropout rates

Outputs:
  oregon_counties_enriched.geojson

Usage:
  python3 scripts/process_data.py  (run from repo root; requires openpyxl)
"""

import json
import re
from pathlib import Path
import openpyxl

BASE = Path(__file__).parent.parent / 'data'

# ── helpers ────────────────────────────────────────────────────────────────

def safe_float(v):
    """Return float or None for suppressed/missing values."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ('*', '', '-', 'N/A', 'n/a'):
        return None
    if s.startswith(('<', '>')):
        return None  # boundary-suppressed (e.g. "<5%", ">95%")
    try:
        return float(s)
    except ValueError:
        return None

def safe_int(v):
    f = safe_float(v)
    return int(round(f)) if f is not None else None

def county_key(name):
    """Normalize to lowercase plain name: 'Baker County' or 'Baker' → 'baker'."""
    cleaned = re.sub(r'\s+county$', '', str(name).strip(), flags=re.IGNORECASE)
    return cleaned.lower()


# ── 1. County boundary polygons ────────────────────────────────────────────

print("Loading county boundary polygons...")
with open(BASE / 'Oregon_Counties_map_20260325.geojson') as f:
    counties_geo = json.load(f)

county_features = {}
for feat in counties_geo['features']:
    p = feat['properties']
    key = county_key(p['instname'])
    county_features[key] = feat

print(f"  {len(county_features)} counties loaded")


# ── 2. Library counts by county and type ──────────────────────────────────

print("Loading library directory...")
with open(BASE / 'Oregon_Library_Directory_20260325.geojson') as f:
    libs_geo = json.load(f)

# Map type_of_library values → output field names
LIB_TYPE_FIELDS = {
    'Public':     'library_public',
    'Academic':   'library_academic',
    'Special':    'library_special',
    'Tribal':     'library_tribal',
    'Volunteer':  'library_volunteer',
    'County law': 'library_county_law',
}
ALL_LIB_TYPE_FIELDS = list(LIB_TYPE_FIELDS.values()) + ['library_other']

lib_data = {}
unmatched_lib_counties = set()

for feat in libs_geo['features']:
    p = feat['properties']
    key = county_key(p.get('county', ''))
    if not key:
        continue

    if key not in lib_data:
        lib_data[key] = {'library_count': 0}
        for field in ALL_LIB_TYPE_FIELDS:
            lib_data[key][field] = 0

    lib_data[key]['library_count'] += 1
    ltype = p.get('type_of_library', '')
    field = LIB_TYPE_FIELDS.get(ltype, 'library_other')
    lib_data[key][field] += 1

    if key not in county_features:
        unmatched_lib_counties.add(key)

total_libs = sum(d['library_count'] for d in lib_data.values())
print(f"  {total_libs} libraries across {len(lib_data)} counties")
if unmatched_lib_counties:
    print(f"  WARNING – library county names not in boundary file: {sorted(unmatched_lib_counties)}")


# ── 3. Graduation rates (4-year cohort, 2024-25) ──────────────────────────

print("Loading graduation rates...")
wb_cohort = openpyxl.load_workbook(
    BASE / 'cohortmediafile2024-2025.xlsx', data_only=True
)
ws_4yr = wb_cohort['4YR State and County']

# Column indices (0-based), confirmed from header row inspection
C_COUNTY     = 0
C_GROUP      = 5
C_COHORT     = 6
C_GRAD_RATE  = 11   # "2024-25 Four-year Cohort Graduation Rate"
C_INST_LVL   = 20   # "Institution Level" → "State Level" | "County"
C_PRIOR_RATE = 21   # "2023-24 Four-year Cohort Graduation Rate"

# Student groups to extract and their output field names
GRAD_GROUPS = {
    'All Students':                               'grad_rate_4yr',
    'American Indian/Alaska Native':              'grad_rate_native',
    'Asian':                                      'grad_rate_asian',
    'Black/African American':                     'grad_rate_black',
    'Hispanic/Latino':                            'grad_rate_hispanic',
    'Multi-Racial':                               'grad_rate_multiracial',
    'Native Hawaiian/Pacific Islander':           'grad_rate_nhpi',
    'White':                                      'grad_rate_white',
    'Students with Disabilities':                 'grad_rate_disabilities',
    'English Learners (Anytime in High School)':  'grad_rate_ell',
    'Male':                                       'grad_rate_male',
    'Female':                                     'grad_rate_female',
}

grad_data = {}  # county_key → {field: value}

for row in ws_4yr.iter_rows(min_row=2, values_only=True):
    if row[C_INST_LVL] != 'County':
        continue
    group = row[C_GROUP]
    if group not in GRAD_GROUPS:
        continue
    key = county_key(row[C_COUNTY])
    if key not in grad_data:
        grad_data[key] = {}

    field = GRAD_GROUPS[group]
    grad_data[key][field] = safe_float(row[C_GRAD_RATE])

    if group == 'All Students':
        grad_data[key]['adjusted_cohort'] = safe_int(row[C_COHORT])
        grad_data[key]['grad_rate_prior'] = safe_float(row[C_PRIOR_RATE])

print(f"  Graduation data for {len(grad_data)} counties")


# ── 4. Dropout rates (2024-25) ─────────────────────────────────────────────

print("Loading dropout rates...")
wb_drop = openpyxl.load_workbook(
    BASE / 'dropouttables2024-2025.xlsx', data_only=True
)
ws_drop = wb_drop['State and County']

# Column indices (0-based)
D_SCHOOL_YR    = 0
D_COUNTY       = 1
D_GROUP        = 2
D_FALL_MEMB    = 3
D_DROP_COUNT   = 4
D_DROP_RATE    = 5
D_INST_TYPE    = 6   # "State" | "County"

dropout_data = {}  # county_key → {dropout_rate, fall_membership}

for row in ws_drop.iter_rows(min_row=2, values_only=True):
    if row[D_INST_TYPE] != 'County':
        continue
    if row[D_GROUP] != 'All Students':
        continue
    key = county_key(row[D_COUNTY])
    dropout_data[key] = {
        'dropout_rate':    safe_float(row[D_DROP_RATE]),
        'fall_membership': safe_int(row[D_FALL_MEMB]),
    }

print(f"  Dropout data for {len(dropout_data)} counties")


# ── 5. Join all datasets onto county boundary features ────────────────────

print("Joining datasets...")

EMPTY_LIB = {'library_count': 0}
for field in ALL_LIB_TYPE_FIELDS:
    EMPTY_LIB[field] = 0

matched_grad = 0
unmatched_grad = []

for key, feat in county_features.items():
    p = feat['properties']
    county_name = re.sub(r'\s+county$', '', p['instname'], flags=re.IGNORECASE)

    libs   = lib_data.get(key, EMPTY_LIB.copy())
    grad   = grad_data.get(key, {})
    drop   = dropout_data.get(key, {})

    new_props = {
        'county_name':      county_name,
        'county_name_full': p['instname'],
        'fips_code':        str(p.get('instcode', '')),
    }
    new_props.update(libs)
    new_props.update(grad)
    new_props.update(drop)

    # Derived metric: libraries per 1,000 enrolled students
    membership = new_props.get('fall_membership')
    lib_count  = new_props.get('library_count', 0)
    if membership and membership > 0:
        new_props['libraries_per_1k_students'] = round(lib_count / membership * 1000, 3)
    else:
        new_props['libraries_per_1k_students'] = None

    feat['properties'] = new_props

    if key in grad_data:
        matched_grad += 1
    else:
        unmatched_grad.append(county_name)

print(f"  Graduation data matched: {matched_grad}/36 counties")
if unmatched_grad:
    print(f"  WARNING – no graduation data for: {sorted(unmatched_grad)}")


# ── 6. Write output GeoJSON ────────────────────────────────────────────────

out_path = BASE / 'oregon_counties_enriched.geojson'
with open(out_path, 'w') as f:
    json.dump(counties_geo, f, separators=(',', ':'))

print(f"\nWrote {out_path}  ({out_path.stat().st_size / 1024:.0f} KB)")


# ── 7. Audit report ────────────────────────────────────────────────────────

print("\n" + "─" * 70)
print("AUDIT REPORT")
print("─" * 70)

feats = counties_geo['features']
print(f"Counties in output:      {len(feats)}")

rates = [f['properties'].get('grad_rate_4yr') for f in feats]
valid_rates = [r for r in rates if r is not None]
null_rates  = [f['properties']['county_name'] for f in feats if f['properties'].get('grad_rate_4yr') is None]
print(f"Counties with grad rate: {len(valid_rates)}/36")
if valid_rates:
    print(f"Grad rate range:         {min(valid_rates):.1f}% – {max(valid_rates):.1f}%")
if null_rates:
    print(f"Suppressed (null) rates: {', '.join(sorted(null_rates))}")

total_lib_out = sum(f['properties'].get('library_count', 0) for f in feats)
print(f"Total libraries in join: {total_lib_out}  (source: {total_libs})")
if total_lib_out != total_libs:
    print(f"  WARNING – library count mismatch!")

print()
print(f"{'County':<20} {'Libs':>5} {'Grad%':>7} {'Drop%':>7} {'Lib/1kStu':>10}")
print("-" * 53)
for feat in sorted(feats, key=lambda f: f['properties']['county_name']):
    props = feat['properties']
    name  = props['county_name']
    libs  = props.get('library_count', 0)
    grad  = f"{props['grad_rate_4yr']:.1f}" if props.get('grad_rate_4yr') is not None else ' N/A'
    drop  = f"{props['dropout_rate']:.2f}"  if props.get('dropout_rate')  is not None else '  N/A'
    lper  = f"{props['libraries_per_1k_students']:.2f}" if props.get('libraries_per_1k_students') is not None else '     N/A'
    print(f"{name:<20} {libs:>5} {grad:>7} {drop:>7} {lper:>10}")

print("─" * 70)
